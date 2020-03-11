
# coding= utf-8
# import sys
# reload(sys)
# sys.setdefaultencoding('utf8')


import os
import io
import six
import json
import time
import datetime
import traceback
from pydub import AudioSegment

from google.cloud import storage
from google.cloud import language
from google.cloud.language import enums
from google.cloud.language import types

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.reader.excel import load_workbook
from tempfile import NamedTemporaryFile


credential_path = "apikey.json"
os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = credential_path

def list_blobs(bucket):
    storage_client = storage.Client()
    blobs = storage_client.list_blobs(bucket)
    print(blobs)
    for blob in blobs:
        file = blob.name
        gcs_url = 'gs://%(bucket)s/%(file)s' % {'bucket':bucket, 'file':file}
        print (gcs_url)
        check_gcs(gcs_url)


def list_blobs1(bucket,prefix):
    storage_client = storage.Client()
    blobs = storage_client.list_blobs(bucket,prefix=prefix)
    print(blobs)
    for blob in blobs:
        file = blob.name
        gcs_url = 'gs://%(bucket)s/%(file)s' % {'bucket':bucket, 'file':file}
        print (gcs_url)
        check_gcs(gcs_url)


def check_gcs(gcs_url):
    # date = '2020-02-09'
    a = (gcs_url.split("/")[-1]) 
    date = a.split("-")[0]+"-"+a.split("-")[1]+"-"+a.split("-")[2]
    # date = datetime.date.today().strftime("%Y-%m-%d")
    with open(date+'.txt','a') as f:
        with open(date+'.txt', "r") as f:
            done=[]
            for line in f.readlines():
                line = line.strip('\n')
                done.append(line)
            if gcs_url in done:
                print('already')
                pass
            else:
                try:
                    transcribe_gcs(gcs_url,date)
                except Exception as err:
                    traceback.print_exc()
                    model_log(__file__,'0:\n%s' % traceback.format_exc(),gcs_url)
                
                

def onefile(bucket,file):
    gcs_url = 'gs://%(bucket)s/%(file)s' % {'bucket':bucket, 'file':file}
    a = (gcs_url.split("/")[-1]) 
    date = a.split("-")[0]+"-"+a.split("-")[1]+"-"+a.split("-")[2]
    print (gcs_url)
    transcribe_gcs(gcs_url,date)


def transcribe_gcs(gcs_url,date):
    from google.cloud import speech
    from google.cloud.speech import enums
    from google.cloud.speech import types
    client = speech.SpeechClient()
    audio = types.RecognitionAudio(uri=gcs_url)
    config = types.RecognitionConfig(
        encoding=enums.RecognitionConfig.AudioEncoding.LINEAR16,
        # sample_rate_hertz=16000,
        enable_automatic_punctuation=True,
        language_code='zh-TW')
    operation = client.long_running_recognize(config, audio)
    response = operation.result(timeout=90)
    toexcel(response,gcs_url,date)


def toexcel(response,gcs_url,date):
    wb = Workbook()
    bn = date+'-1.xlsx'
    # 檢查檔案是否存在
    if os.path.isfile(bn):
        wb = load_workbook(bn)
        checkexcel(response,gcs_url,wb,bn,date)
    else:
        checkexcel(response,gcs_url,wb,bn,date)


def checkexcel(response,gcs_url,wb,bn,date):
    with open(date+'.txt','a') as f:
        f.write(gcs_url+'\n')

    ws = wb.active
    add = {}
    #FIXME:有修正過要把所有依樣day加入
    for result in response.results:
            alternative = result.alternatives[0]
            if gcs_url in add.keys():
                add[gcs_url] = add[gcs_url]+alternative.transcript
            else:
                add[gcs_url] = alternative.transcript
    for key in add.keys():
        ws.append([key,add[key]])        

    wb.save(bn)
    cut(bn,gcs_url)


def cut(bn,gcs_url):
    wb = load_workbook(bn)
    ws = wb.active
    first_column = ws['B']
    for x in range(len(first_column)): 
        # print(first_column[x].value) 
        clean_text(first_column[x].value,gcs_url)
    os.remove(bn)

def syntax_text(text,gcs_url):
    text = text
    client = language.LanguageServiceClient()

    if isinstance(text, six.binary_type):
        text = text.decode('utf-8')
    document = types.Document(
        content=text,
        language = 'zh-Hant',
        type=enums.Document.Type.PLAIN_TEXT)

    tokens = client.analyze_syntax(document).tokens
    final = {}
    for token in tokens:
        part_of_speech_tag = enums.PartOfSpeech.Tag(token.part_of_speech.tag)
    
        if part_of_speech_tag.name in final.keys():
            final[part_of_speech_tag.name].append(token.text.content)
        else:
            final[part_of_speech_tag.name] = [token.text.content]
    
    if not 'VERB' in final.keys():
        final['VERB'] = ''
    if not 'NOUN' in final.keys():
        final['NOUN'] = ''
    if not 'ADJ' in final.keys():
        final['ADJ'] = ''

    # final['VERB'] = json.dumps(final['VERB'], encoding='UTF-8', ensure_ascii=False)
    # final['NOUN'] = json.dumps(final['NOUN'], encoding='UTF-8', ensure_ascii=False)
    # final['ADJ'] = json.dumps(final['ADJ'], encoding='UTF-8', ensure_ascii=False)

    return final['VERB'],final['NOUN'],final['ADJ']


def sentiment_text(text,gcs_url):
    text = text
    client = language.LanguageServiceClient()
    try:
        text = text.decode('utf-8')
    except AttributeError:
        pass

    document = types.Document(
        content=text,
        language = 'zh-Hant',
        type=enums.Document.Type.PLAIN_TEXT)

    sentiment = client.analyze_sentiment(document).document_sentiment
    return sentiment.score,sentiment.magnitude
    # print('Score: {}'.format(sentiment.score))
    # print('Magnitude: {}'.format(sentiment.magnitude))


def clean_text(text,gcs_url):
    have_1 = ["零", "壹", "貳", "參", "肆", "伍", "陸", "柒", "捌", "玖"]
    try:
        for word in text[:10]:
            if word in have_1:
                text = text.replace(word, str(have_1.index(word)),1)
    except TypeError:
        return text

    have_2 = ['零','一','二','三','四','五','六','七','八','九']
    for word in text[:10]:
        if word in have_2:
            text = text.replace(word, str(have_2.index(word)),1)
    if '外科' in text:
        text = text.replace('外科', '外客')
    text = text.strip()
    # print(text)
    find_name(text,gcs_url)


def find_name(text,gcs_url):
    test = ['晚餐', '午餐','早餐', '下午茶']
    temp_text = text[:15]
    word = next((x for x in test if x in temp_text), None)

    if word:
        # print(word)
        if word in text[:15]:
            
            time_def(text,word,gcs_url)
        else:
            time = ''
            name = ''
            part = ''
        
            other = text 
            score,magnitude = sentiment_text(text,gcs_url)
            verb, noun, adj = syntax_text(text,gcs_url)
            wirte_excel(part,time,name,other,verb, noun, adj,score,magnitude,gcs_url)

    else:
        test = ['房號', '房客','外客']
        word = next((x for x in test if x in text), None)
        if word:
            if word in text[:15]:
                time_def(text,word,gcs_url)
        else:
                time = ''
                name = ''
                part = ''

                other = text 
                score,magnitude = sentiment_text(text,gcs_url)
                verb, noun, adj = syntax_text(text,gcs_url)
                wirte_excel(part,time,name,other,verb, noun, adj,score,magnitude,gcs_url)


def time_def(text,cut,gcs_url):
    time = text.split(cut,1)[0]
    part = ''
    other = cut + text.split(cut,1)[1] 
    # print(time)
    if time == '':
        name,other = cut_name(other,gcs_url) 
        final_time = ''
    
        score,magnitude = sentiment_text(other,gcs_url)
        verb, noun, adj = syntax_text(other,gcs_url)
    
        return wirte_excel(part,final_time,name,other,verb, noun, adj,score,magnitude,gcs_url)

    for word in time:
        if word.isdigit():
    
            digit = time.index(word)
            
            break
        else:
            digit = None
    if digit:
        part = time[0:digit]
        time = time[digit:]
    # print(part)
        
    have = ['月','號','日','早上','早','下午','下','午','晚上','點','分','半','晚','上']

    for word in time:
        if word not in have and word.isdigit() != True:
            time = time.replace(word, '')
            # print(word)
            
    # print(time)
    if time == '':
        name,other = cut_name(other,gcs_url) 
        final_time = ''
    
        score,magnitude = sentiment_text(other,gcs_url)
        verb, noun, adj = syntax_text(other,gcs_url)
    
        return wirte_excel(part,final_time,name,other,verb, noun, adj,score,magnitude,gcs_url)

    today = datetime.datetime.now() 
    time = str(today.year)+'年'+time

    if '日' in time:
        time = time.replace('日', '號')
    
    if '點' not in time:
        final_time = cut_time_day(time,gcs_url)
    
    else:
    
        if '月' and '號' in time:
            if '早上' in time:
                time = time.replace('早上', '')
            elif '下午' or '晚上' in time:
                if '下午' in time:
                    new = time[:(time.index('下午')+2)]+str(int(time[(time.index('下午')+2):time.index('點')]) +12)+time[time.index('點'):]
                    time = new.replace('下午', '')
                    # print(time)    
                else:
                    new = time[:(time.index('晚上')+2)]+str(int(time[(time.index('晚上')+2):time.index('點')]) +12)+time[time.index('點'):]
                    time = new.replace('晚上', '')
            else:
                time = time
                
        else:
            if '早上' in time:
                temp = list(time)
                station = temp.index('早') 
                time =  "".join(temp[0:(station)])+'號'+"".join(temp[station:]) 
                time = time.replace('早上', '')
            elif '下午' or '晚上' in time:
                if '下午' in time:
                    temp = list(time)
                    station = temp.index('下') 
                    time =  "".join(temp[0:(station)])+'號'+"".join(temp[station:]) 
                    new = time[:(time.index('下午')+2)]+str(int(time[(time.index('下午')+2):time.index('點')]) +12)+time[time.index('點'):]
                    time = new.replace('下午', '')
                    # print(time)
                
                else:
                    temp = list(time)
                    station = temp.index('晚') 
                    time =  "".join(temp[0:(station)])+'號'+"".join(temp[station:]) 
                    new = time[:(time.index('晚上')+2)]+str(int(time[(time.index('晚上')+2):time.index('點')]) +12)+time[time.index('點'):]
                    time = new.replace('晚上', '')
        # print(time)
        final_time = cut_time_hour(time,gcs_url)

    name,other = cut_name(other,gcs_url) 
    score,magnitude = sentiment_text(other,gcs_url)
    verb, noun, adj = syntax_text(other,gcs_url)

    wirte_excel(part,final_time,name,other,verb, noun, adj,score,magnitude,gcs_url)


def cut_time_day(text,gcs_url):
    if '號' in text:
        pass
    else:
        text = text + '號'
    try:
        time = datetime.datetime.strptime(text, '%Y年%m月%d號').date()
    except ValueError:
        return text
    # print(time)
    return time


def  cut_time_hour(text,gcs_url):
    if '半' in text:
        try:
            text = text.split('半',1)[0]
            time = datetime.datetime.strptime(text, '%Y年%m月%d號%H點')
            time +=  datetime.timedelta(minutes=30)
        except ValueError:
            return text
    elif '分' in text:
        try:
            time = datetime.datetime.strptime(text, '%Y年%m月%d號%H點%M分')
        except ValueError:
            return text
    else:
        try:
            test = list(text)
            station = text.index('點')
            if '點' == test[-1]:
                time = datetime.datetime.strptime(text, '%Y年%m月%d號%H點')
            elif test[station+1].isdigit() :
                test.append('分')
                test = "".join(test)
                time = datetime.datetime.strptime(test, '%Y年%m月%d號%H點%M分')
        except ValueError:
            return text

    return time
    # print(time)


def cut_name(text,gcs_url):
    test = ['小姐','先生','客人','夫妻']

    word = next((x for x in test if x in text), None)
    if word :
        name = text.split(word,1)[0]+ word
        other = text.split(word,1)[1]# 指切前面的就好 
    else:
        name,other = '', text

    return name,other


def wirte_excel(part,time,name,other,verb, noun, adj,score,magnitude,gcs_url):
    # date = '2020-02-09'
    a = (gcs_url.split("/")[-1]) 
    date = a.split("-")[0]+"-"+a.split("-")[1]+"-"+a.split("-")[2]
    # date = datetime.date.today().strftime("%Y-%m-%d")
    wb = Workbook()
    fn = date+'.xlsx'
    # 檢查檔案是否存在
    if os.path.isfile(fn):
        wb = load_workbook(fn)
        finalexcel(wb,fn,part,time,name,other,verb, noun, adj,score,magnitude,gcs_url)
    else:
        finalexcel(wb,fn,part,time,name,other,verb, noun, adj,score,magnitude,gcs_url)

def finalexcel(wb,fn,part,time,name,other,verb, noun, adj,score,magnitude,gcs_url):
    ws = wb.active
    ws['A1'] = 'gcs_url'   
    ws['B1'] = 'part'
    ws['C1'] = 'time'
    ws['D1'] = 'name'
    ws['E1'] = 'event'
    ws['F1'] = 'verb'
    ws['G1'] = 'noun'
    ws['H1'] = 'adj'
    ws['I1'] = 'score'
    ws['J1'] = 'magnitude'
    
    if score >0.1:
        react = '正面'
    elif score == 0:
        react = '中立'
    else:
        react = '負面'

    ws.append([gcs_url,part,str(time),name,str(other),str(verb),str(noun),str(adj) ,score,magnitude,react])    
    wb.save(fn)
    print('success')
    tostorage(fn)

def tostorage(xlsx):
    gcs = storage.Client()
    bucket = gcs.get_bucket('lalu-aud-tw')
    folder = 'data'
    filename = "%s/%s" % (folder, xlsx)
    blob = bucket.blob(filename)
    blob.upload_from_filename(xlsx)
    print('save to storage')


def upload(bucket,filename):
    client = storage.Client()
    bucket1 = client.get_bucket(bucket)
    blob = bucket1.blob(filename)
    blob.upload_from_filename(filename)
    file = filename
    gcs_url = 'gs://%(bucket)s/%(file)s' % {'bucket':bucket, 'file':file}
    print(gcs_url)
    transcribe_gcs(gcs_url)


def model_log(file_path,exe_status,gcs_url):
    record_time = time.strftime("%Y%m%d %H:%M:%S")
    log_file = open(f'{record_time[0:8]}.log',mode='a',encoding='utf-8')
    log_file.write('執行程式:'+str(file_path)+'\n')
    log_file.write('執行時間:'+record_time+'\n')
    log_file.write('處理中的音檔:'+str(gcs_url)+'\n')
    log_file.write('執行狀態:'+str(exe_status)+'\n')
    log_file.write('---------------------------------------'+'\n')

print('start')
bucket='lalu-aud-tw'
list_blobs1(bucket, datetime.date.today())
print('end')
# list_blobs1(bucket, '2020-02-28')
# tostorage('2020-02-28.xlsx')